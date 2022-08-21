import openpyxl

class Funexcel():
    
    def __init__(self,driver):
        self.driver = driver
    
    def obtener_Cant_Filas(self,path,sheetName):
        Worbook = openpyxl.load_workbook(path)
        sheet = Worbook[sheetName]
        return (sheet.max_row)
    
    def obtener_Cant_Columnas(self,path,sheetName):
        Worbook = openpyxl.load_workbook(path)
        sheet = Worbook[sheetName]
        return (sheet.max_column)
    
    def lectura_Datos(self,path,sheetName,fila,columna):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return sheet.cell(row=fila,column=columna).value
    
    def escritura_Datos(self,path,sheetName,fila,columna,data):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        sheet.cell(row=fila,column=columna).value = data
        Workbook.save(path)
    